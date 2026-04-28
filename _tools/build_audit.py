#!/usr/bin/env python3
"""Build seo-audit.xlsx from _crawl/inventory.json."""
import json
from collections import Counter, defaultdict
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
INV = json.loads((ROOT / "_crawl" / "inventory.json").read_text())
OUT = ROOT / "seo-audit.xlsx"

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
BODY_FONT = Font(name="Arial", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(border_style="thin", color="DDDDDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SEVERITY_FILL = {
    "HIGH": PatternFill("solid", start_color="F8CBAD"),
    "MED":  PatternFill("solid", start_color="FFE699"),
    "LOW":  PatternFill("solid", start_color="C6E0B4"),
}

def derive_issues(p):
    out = []
    if not p.get("title"): out.append(("MISSING_TITLE", "HIGH", "Title tag empty"))
    elif p["titleLength"] < 30: out.append(("TITLE_TOO_SHORT", "MED", f"Title is {p['titleLength']} chars (target 30-60)"))
    elif p["titleLength"] > 60: out.append(("TITLE_TOO_LONG", "MED", f"Title is {p['titleLength']} chars (target 30-60)"))
    if not p.get("description"): out.append(("MISSING_DESC", "HIGH", "Meta description empty"))
    elif p["descriptionLength"] < 70: out.append(("DESC_TOO_SHORT", "MED", f"Description is {p['descriptionLength']} chars (target 70-160)"))
    elif p["descriptionLength"] > 160: out.append(("DESC_TOO_LONG", "MED", f"Description is {p['descriptionLength']} chars (target 70-160)"))
    if p.get("h1Count", 0) == 0: out.append(("NO_H1", "HIGH", "Page has no H1"))
    if p.get("h1Count", 0) > 1: out.append(("MULTIPLE_H1", "MED", f"{p['h1Count']} H1s on page"))
    if not p.get("canonical"): out.append(("NO_CANONICAL", "MED", "No canonical URL"))
    if p.get("imgsMissingAlt", 0) > 0: out.append(("IMGS_MISSING_ALT", "MED", f"{p['imgsMissingAlt']} image(s) missing alt"))
    if p.get("robots") and "noindex" in p["robots"].lower(): out.append(("NOINDEX", "LOW", "Page is set to noindex"))
    if p.get("wordCount", 0) < 300: out.append(("THIN_CONTENT", "MED", f"Only {p['wordCount']} words"))
    if p.get("jsonLdCount", 0) == 0: out.append(("NO_SCHEMA", "MED", "No JSON-LD schema present"))
    if not p.get("ogTitle"): out.append(("NO_OG_TITLE", "LOW", "Missing og:title"))
    if not p.get("ogImage"): out.append(("NO_OG_IMAGE", "LOW", "Missing og:image"))
    return out

def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

wb = Workbook()

# ========== Pages sheet ==========
ws = wb.active
ws.title = "Pages"
cols = [
    ("URL", 38), ("Slug", 22), ("Status", 8),
    ("Title", 50), ("Title Len", 10),
    ("Description", 50), ("Desc Len", 10),
    ("Canonical", 38), ("Robots Meta", 18),
    ("H1 Count", 9), ("H1 Text", 40),
    ("Word Count", 11), ("Img Count", 10), ("Imgs Missing Alt", 16),
    ("Forms", 7), ("JSON-LD Count", 14),
    ("Internal Links", 14), ("External Links", 14),
    ("OG Title", 30), ("OG Image", 30),
    ("Issues", 50),
]
ws.append([c[0] for c in cols])
for p in INV:
    issues = derive_issues(p)
    ws.append([
        p.get("url", ""), p.get("slug", ""), p.get("status", ""),
        p.get("title", ""), p.get("titleLength", 0),
        p.get("description", ""), p.get("descriptionLength", 0),
        p.get("canonical", ""), p.get("robots", ""),
        p.get("h1Count", 0), " | ".join(p.get("h1s", []))[:200],
        p.get("wordCount", 0), p.get("imgCount", 0), p.get("imgsMissingAlt", 0),
        p.get("forms", 0), p.get("jsonLdCount", 0),
        p.get("internalLinks", 0), p.get("externalLinks", 0),
        p.get("ogTitle", "")[:100], p.get("ogImage", "")[:100],
        ", ".join(i[0] for i in issues),
    ])
style_header(ws, len(cols))
autosize(ws, [c[1] for c in cols])
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.font = BODY_FONT
        cell.alignment = WRAP
        cell.border = BORDER

# ========== Issues sheet (one row per page-issue) ==========
ws2 = wb.create_sheet("Issues")
icols = [("URL", 38), ("Slug", 22), ("Issue", 22), ("Severity", 10), ("Detail", 60), ("Recommended Fix", 60)]
ws2.append([c[0] for c in icols])

FIXES = {
    "MISSING_TITLE": "Add a unique <title> 30-60 chars including primary keyword + brand.",
    "TITLE_TOO_SHORT": "Expand title to 30-60 chars; include keyword and brand.",
    "TITLE_TOO_LONG": "Trim title to <=60 chars to avoid SERP truncation.",
    "MISSING_DESC": "Write a 70-160 char meta description with a clear CTA.",
    "DESC_TOO_SHORT": "Expand meta description to 70-160 chars.",
    "DESC_TOO_LONG": "Trim meta description to <=160 chars.",
    "NO_H1": "Add a single H1 that matches user intent.",
    "MULTIPLE_H1": "Demote extra H1s to H2/H3; keep one H1 per page.",
    "NO_CANONICAL": "Add <link rel='canonical'> pointing to the production URL.",
    "IMGS_MISSING_ALT": "Add descriptive alt text on every <img> (decorative -> alt='').",
    "NOINDEX": "Confirm noindex is intentional (staging OK; production thank-you/legal pages typically OK).",
    "THIN_CONTENT": "Expand content >=300 words or noindex if utility page.",
    "NO_SCHEMA": "Add LegalService/LocalBusiness/FAQPage JSON-LD as appropriate.",
    "NO_OG_TITLE": "Add og:title for social previews.",
    "NO_OG_IMAGE": "Add og:image (1200x630) for social previews.",
}

for p in INV:
    for code, sev, detail in derive_issues(p):
        ws2.append([p.get("url", ""), p.get("slug", ""), code, sev, detail, FIXES.get(code, "")])
style_header(ws2, len(icols))
autosize(ws2, [c[1] for c in icols])
for row in ws2.iter_rows(min_row=2):
    sev_cell = row[3]
    if sev_cell.value in SEVERITY_FILL:
        sev_cell.fill = SEVERITY_FILL[sev_cell.value]
        sev_cell.font = Font(name="Arial", bold=True, size=10)
    for cell in row:
        if cell.column != 4:
            cell.font = BODY_FONT
        cell.alignment = WRAP
        cell.border = BORDER

# ========== Assets sheet ==========
ws3 = wb.create_sheet("Assets")
acols = [("Image Src", 70), ("Pages Using", 8), ("Has Alt On All", 16), ("Sample Alt", 40)]
ws3.append([c[0] for c in acols])
img_pages = defaultdict(list)
img_alts = defaultdict(list)
for p in INV:
    for im in p.get("imgs", []):
        img_pages[im["src"]].append(p["slug"])
        img_alts[im["src"]].append(im.get("alt", ""))
for src, pages in sorted(img_pages.items()):
    alts = img_alts[src]
    has_all = "Yes" if all(a for a in alts) else "No"
    sample = next((a for a in alts if a), "")
    ws3.append([src, len(pages), has_all, sample])
style_header(ws3, len(acols))
autosize(ws3, [c[1] for c in acols])
for row in ws3.iter_rows(min_row=2):
    for cell in row:
        cell.font = BODY_FONT
        cell.alignment = WRAP
        cell.border = BORDER

# ========== Summary sheet ==========
ws4 = wb.create_sheet("Summary", 0)
ws4.append(["AccidentLogic SEO Audit"])
ws4["A1"].font = Font(name="Arial", bold=True, size=16)
ws4.append([])
ws4.append(["Source", "https://accidentlogic.com/"])
ws4.append(["Pages crawled", len(INV)])
ws4.append(["Total issues", "=SUM(C8:C20)"])
ws4.append([])
ws4.append(["Issue", "Severity", "Count"])
counts = Counter()
sev_for = {}
for p in INV:
    for code, sev, _ in derive_issues(p):
        counts[code] += 1
        sev_for[code] = sev
for code, n in sorted(counts.items(), key=lambda x: -x[1]):
    ws4.append([code, sev_for[code], n])
# style summary
for c in range(1, 4):
    cell = ws4.cell(row=7, column=c)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
ws4.freeze_panes = "A8"
autosize(ws4, [28, 12, 10])
for row in ws4.iter_rows(min_row=8, max_row=7 + len(counts)):
    sev_cell = row[1]
    if sev_cell.value in SEVERITY_FILL:
        sev_cell.fill = SEVERITY_FILL[sev_cell.value]
    for cell in row:
        if not cell.font.bold:
            cell.font = BODY_FONT
        cell.border = BORDER

wb.save(OUT)
print(f"Wrote {OUT}")
